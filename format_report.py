from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break, RowBreak, ColBreak


def format(file, job_times, formated_date):

    workbook = load_workbook(file)
    
    sheet = workbook.active 
    
    row_count = sheet.max_row
    column_count = sheet.max_column
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.paperSize   = sheet.PAPERSIZE_LEGAL
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_margins           = PageMargins(left=0.6, right=0.6, top=1.9, bottom=1.9, header=0.5, footer=0.8)

    # Set the text for the center of the odd page header
    sheet.oddHeader.center.text  = 'GO and UP Crew Lineup'
    sheet.oddHeader.center.size  = 22
    sheet.oddHeader.center.font  = 'Arial'
    sheet.oddHeader.center.color = '000000'

    # Set a left-aligned header with page numbers
    sheet.oddHeader.right.text = formated_date
    sheet.oddHeader.right.size = 18
    sheet.oddHeader.right.font = 'Arial'
    sheet.oddHeader.right.color ='000000'

    row_count = sheet.max_row
    column_count = sheet.max_column
    sheet.sheet_view.showGridLines = False

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'),   bottom=Side(style='thin'))

    # Fonts
    header_font  = Font(bold=True, name='Arial',size=12)
    page_font    = Font(name='Arial', size=12)
    vacancy_font = Font(color = 'FF0000', bold=True, name='Arial', size=12)
    
    # Background colours
    newly_qualified_green = PatternFill(start_color='C6E0B4', fill_type='solid')
    vacancy_yellow        = PatternFill(start_color='FFFF00', fill_type='solid')
    light_blue_fill       = PatternFill(start_color='BDD7EE', fill_type='solid')
    not_important_grey    = PatternFill(start_color='C0C0C0', fill_type='solid')
    


    # Set up cells    
    for row_num in range(1, row_count+1):
        for col_num in range(1, column_count+1):
            cell            = sheet.cell(row=row_num, column=col_num)
            cell.font       = page_font
            cell.alignment  = Alignment(horizontal='left')
            cell.border     = thin_border
            
    
    active_columns     = 'ABCDEFG' 
    not_active_columns = 'HIJKLMN'
    
    # format header
    for col in active_columns:
        sheet[f'{col}1'].font = header_font
        sheet[f'{col}1'].fill = light_blue_fill
        sheet[f'{col}1'].alignment = Alignment(horizontal='left')
    for col in not_active_columns:
        sheet[f'{col}1'].fill = not_important_grey

        

    # set up column widths
    sheet.column_dimensions['A'].width = 9.89
    sheet.column_dimensions['B'].width = 9.89
    sheet.column_dimensions['C'].width = 45
    sheet.column_dimensions['D'].width = 45
    sheet.column_dimensions['E'].width = 45
    sheet.column_dimensions['F'].width = 45
    sheet.column_dimensions['G'].width = 45
    sheet.column_dimensions['H'].width = 12.89
    sheet.column_dimensions['I'].width = 12.89
    sheet.column_dimensions['J'].width= 41.67
    sheet.column_dimensions['K'].width= 41.67
    sheet.column_dimensions['L'].width= 41.67
    sheet.column_dimensions['M'].width= 41.67
    sheet.column_dimensions['N'].width= 41.67
    
    #sheet.col_breaks.append(Break(id=8))
    sheet.col_breaks = ColBreak(brk=(Break(id=7),))
    sheet.print_area = f'A1:G{row_count}'
    
    
    for row_num in range (2, row_count + 1):
        Tour_number = str(sheet.cell(row=row_num, column = 1).value // 10)
        QCTO_cell = sheet.cell(row=row_num, column = 3)
        CTO_cell  = sheet.cell(row=row_num, column = 4)
        CSA_cell  = sheet.cell(row=row_num, column = 5)
        
        if Tour_number in job_times:
            if  QCTO_cell.value is None:
                QCTO_cell.value = job_times[Tour_number]
                QCTO_cell.fill  = vacancy_yellow
                QCTO_cell.font  = vacancy_font
            if QCTO_cell.value is not None and 'th)' in QCTO_cell.value:
                QCTO_cell.fill  = newly_qualified_green

            if  CTO_cell.value is None:
                CTO_cell.value = job_times[Tour_number]
                CTO_cell.fill  = vacancy_yellow
                CTO_cell.font  = vacancy_font
            if CTO_cell.value is not None and 'th)' in CTO_cell.value:
                CTO_cell.fill  = newly_qualified_green


            if  Tour_number < '50000' and CSA_cell.value is None:
                CSA_cell.value = job_times[Tour_number]
                CSA_cell.fill  =  vacancy_yellow
                CSA_cell.font  = vacancy_font

        if Tour_number >= '900' and Tour_number < '999':
            QCTO_cell.fill = light_blue_fill
            CTO_cell.fill  = light_blue_fill
            if CSA_cell.value is None:
                CSA_cell.fill  = vacancy_yellow
                CSA_cell.font  = vacancy_font

        
        if int(Tour_number) >= 50000 and CSA_cell.value is None:
            CSA_cell.fill = light_blue_fill


   # Save the modified workbook
    workbook.save(file)     

